import numpy as np
import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import FORMULAE
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
os.chdir('/home/saul/Business')

def test():
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "hello"
    sheet["B1"] = "world!"
    workbook.save(filename="hello_world.xlsx")

def excel_work():
     workbook = load_workbook(filename="sample.xlsx")
     worksheet = workbook['Sheet2']
     #sheet = workbook.active
     value = worksheet["C2"].value
     row_count = worksheet.max_row
     column_count = worksheet.max_column
     print("Row count ", row_count)
     print("Column ", column_count)

     #for row in worksheet.iter_rows(values_only=True):
         #print(row)
         
     worksheet.cell(row=33, column=2).value = 'https://www.softwaretestinghelp.com/python-openpyxl-tutorial/'
     worksheet.cell(row=34, column=2).value = 'https://openpyxl.readthedocs.io/en/latest/pandas.html'
     worksheet.insert_rows(13)
     workbook.save("sample.xlsx")
     print("Saved")
   
     #for row in worksheet.rows:
     #     print (row)

def get_function_ranges(cellFormula):
    print("Cell formula ", cellFormula)
    formula = cellFormula.translate({ord(char): None for char in '='})
    funct = formula.split("(")[0]
    print("Function ", funct)
    ranges = formula[formula.find("(")+1:formula.find(")")]
    print("Ranges ", ranges)
    start = ranges.split(':')[0]
    end = ranges.split(':')[1]
    return funct, start, end

def update_formulas(worksheet, funct, row_shift, start_row, end_row):
    #new_start_row = row_shift + int(start_row)
    new_start_row  = 5
    new_end_row = row_shift + int(end_row)
    column = "C"
    worksheet["C3"] = '=' + funct + '(' + column + str(new_start_row) + ':' + column + str(new_end_row) + ')'
    word = '=' + funct + '(' + column + str(new_start_row) + ':' + column + str(new_end_row) + ')'
    print("Word ", word)

def count_dates(worksheet):
    print("Count Dates called !!")
    row_count = worksheet.max_row
    column_count = worksheet.max_column
    row_location = 3
    print("Max Row count ", row_count)
    print("Max Column ", column_count)
    # insert insert three roe at row three
    row_shift_pos = 5
    row_shift = 3

    for row in range(row_shift):
        worksheet.insert_rows(row + row_shift_pos)
  
    print("Row ", worksheet.max_column +1)
    for traverse_row in range(row_location, row_location + 1): 
        for column in "CDEFGHJI":  #Here you can add or reduce the columns
            cell_name = "{}{}".format(column, traverse_row)
            print("Cell Name ", cell_name)
            print("Col values ", worksheet[cell_name].value) # the value of the specific cell
            # if there is new rows inserted

            if row_shift > 0:
                print('{} rows inserted'.format(row_shift))
                funct, start, end = get_function_ranges(worksheet[cell_name].value)
                start_row = re.findall("\d+", start)[0]
                #print("Start Row ", start_row)
                end_row = re.findall("\d+", end)[0]
                #print("End Row ", end_row)
                shift_range = int(end_row) - int(start_row) + 1
                update_formulas(worksheet, funct, row_shift, start_row, end_row)

def copy_formula():
    #print("Copy Formula Called")
    workbook = load_workbook(filename="sample_v2.xlsx")
    worksheet = workbook['Sheet2']
    count_dates(worksheet)
    #funct, start, end = get_function_ranges(worksheet["C3"].value)
    #print("Function {} start at {} and ends at {}".format(funct, start, end))
    #start_row = re.findall("\d+", start)[0]
    #shift_range = int(end_row) - int(start_row) + 1
    #sheet = workbook.active
    value = worksheet["C2"].value
    #print(value)
    row_count = worksheet.max_row
    column_count = worksheet.max_column
    # insert insert three roe at row three
    row_shift_pos = 5
    row_shift = 3
 
    #for row in range(row_shift):
        #worksheet.insert_rows(row + row_shift_pos)
    #update_formulas(worksheet, funct, row_shift, start_row, end_row)
   
    #Add formula 
    worksheet["D7"] = "=AVERAGE(D9:D26)"
    workbook.save("sample_v3.xlsx")

if __name__ == '__main__':
    #test()
    #excel_work()
    copy_formula()
