import numpy as np
import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import FORMULAE
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
os.chdir('/home/saul/Business')

def test():
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "hello"
    sheet["B1"] = "world!"
    workbook.save(filename="hello_world.xlsx")

def excel_work():
     workbook = load_workbook(filename="sample.xlsx")
     worksheet = workbook['Sheet2']
     #sheet = workbook.active
     value = worksheet["C2"].value
     row_count = worksheet.max_row
     column_count = worksheet.max_column
     print("Row count ", row_count)
     print("Column ", column_count)
     #for row in worksheet.iter_rows(values_only=True):
         #print(row)
 
     worksheet.cell(row=33, column=2).value = 'World Wide'
     worksheet.cell(row=34, column=2).value = 'Text Add'
     worksheet.insert_rows(13)
     workbook.save("sample.xlsx")

def get_function_ranges(cellFormula):
    print("Cell formula ", cellFormula)
    formula = cellFormula.translate({ord(char): None for char in '='})
    funct = formula.split("(")[0]
    print("Function ", funct)
    ranges = formula[formula.find("(")+1:formula.find(")")]
    print("Ranges ", ranges)
    start = ranges.split(':')[0]
    print("Start !!! ", start)
    end = ranges.split(':')[1]
    return funct, ranges, start, end

def update_formulas(workbook, worksheet, column_label, column, funct, row_shift, start_row, end_row):
    print("Update Formulas called !!")
    row_location = 3
    #new_start_row = row_shift + int(start_row)
    print("Column ", column)
    print("Column Labels ", column_label)
    new_start_row  = row_shift + int(start_row)
    new_end_row = row_shift + int(end_row)
    word = '=' + funct + '(' + column_label + str(new_start_row) + ':' + column_label +  str(new_end_row) + ')'
    print("Word ", word)
    print("Start Row ", start_row)
    print("Column ", column)
    print("Word ", word, '\n')
    #worksheet[column_coordinate]= '=' + funct + '(' + column + str(new_start_row) + ':' + column + str(new_end_row) + ')'
    worksheet.cell(row= row_location, column=column).value = word
    workbook.save("sample_v3.xlsx")

def count_dates(workbook, worksheet):
    print("Count Dates called !!")
    row_count = worksheet.max_row
    column_count = worksheet.max_column
    row_location = 3
    column = 3
    print("Max Row count ", row_count)
    print("Max Column ", column_count)
    # insert insert three roe at row three
    row_shift_pos = 5
    row_shifts = 3
  
    for row in range(row_shifts):
        worksheet.insert_rows(row + row_shift_pos)
    #print("Row ", worksheet.max_column +1)
    for traverse_row in range(row_location, row_location + 1): 
        while column <= worksheet.max_column:
            cell_name = "{}{}".format(column, traverse_row)
            #print("Cell Name ", cell_name)
            #print("Col values ", worksheet.cell(row=traverse_row, column=column).value) # the value of the specific cell
            # if there is new rows inserted

            if row_shifts > 0:
                funct, ranges, start, end = get_function_ranges(worksheet.cell(row=traverse_row, column=column).value )
                #get integer part
                start_row = re.findall("\d+", start)[0]
                print("Start Row ", start_row)
                end_row = re.findall("\d+", end)[0]
                print("End Row ", end_row)
                column_label = re.findall("\w", start)[0]
                print("Column Label ", column_label)
                shift_range = int(end_row) - int(start_row) + 1
                update_formulas(workbook, worksheet, column_label, column, funct, row_shifts, start_row, end_row)
            column += 1
        
def copy_formula():
    #print("Copy Formula Called")
    workbook = load_workbook(filename="sample_v2.xlsx")
    worksheet = workbook['Sheet2']
    count_dates(workbook, worksheet)
    #sheet = workbook.active
    value = worksheet["C2"].value
    row_count = worksheet.max_row
    column_count = worksheet.max_column
    #print("Row count ", row_count)
    #print("Column ", column_count)
    
    #Add formula 
    worksheet["D7"] = "=AVERAGE(D9:D26)"
    workbook.save("sample_v3.xlsx")

if __name__ == '__main__':
    #excel_work()
    copy_formula()
